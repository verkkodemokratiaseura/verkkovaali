# <OpenID voting platform> Copyright (c) 2012  Suomen Verkkodemokratiaseura ry  <toimisto@verkkodemokratia.fi> This program is free software: you can redistribute it and/or modify it under the terms of the GNU Affero General Public License as published by the Free Software Foundation, either version 3 of the License, or (at your option) any later version. This program is distributed in the hope that it will be useful, but WITHOUT ANY WARRANTY; without even the implied warranty of MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the file LICENSE.TXT in the root directory of this program and GNU Affero General Public License for more details. You should have received a copy of the GNU Affero General Public License along with this program.  If not, see <http://www.gnu.org/licenses/>.
# -*- coding: utf-8 -*-
"""
Functionality for the school representatives to interact with the system.
"""
from cStringIO import StringIO
from datetime import date
from datetime import datetime
from nuorisovaalit.models import Candidate
from nuorisovaalit.models import Voter
from nuorisovaalit.models import VotingLog
from nuorisovaalit.results import dhont_selection
from nuorisovaalitadmin.models import CSVSubmission
from nuorisovaalitadmin.models import DBSession
from nuorisovaalitadmin.views import disable_caching
from nuorisovaalitadmin.views import disable_caching_explorer
from nuorisovaalitadmin.views.login import authenticated_user
from openpyxl.reader.excel import load_workbook
from pyramid.exceptions import Forbidden
from pyramid.url import route_url
from webob import Response
from webob.exc import HTTPFound

import csv
import logging
import os
import re
import tempfile
import xlrd
import xlwt


#: Regular expression to match a date (DD.MM.YYYY).
RE_DOB = re.compile(ur'^(\d{1,2})\.(\d{1,2})\.(\d{4})$')
#: Regular expression to match a short date.
RE_DOB_SHORT = (
    # 151179
    re.compile(ur'^(\d{2})(\d{2})(\d{2})$'),
    # 15.11.79
    re.compile(ur'^(\d{1,2})\.(\d{1,2})\.(\d{2})$'),
    # 151179-666X, 1511796667
    re.compile(ur'^(\d{2})(\d{2})(\d{2})-?\d{3}[a-z0-9]$', re.I))
#: Regular expression to match an email address.
RE_EMAIL = re.compile(ur'^[-a-z0-9_.]+@(?:[-a-z0-9]+\.)+[a-z]{2,6}$', re.IGNORECASE)
#: Regular expression to match a normalized (whitespace and hyphens removed) GSM number.
RE_GSM = re.compile(ur'^((00|\+?)358|0)[1-9]{1}[0-9]{6,}$')


def csv_response(rows, encoding='ISO-8859-1', filename='response.csv'):
    """Create a downloadable webob.Response instance with the given
    row data as a CSV file.
    """
    stream = StringIO()
    writer = csv.writer(stream, dialect='excel')
    for row in rows:
        # Encode all columns in the row since the csv.writer cannot
        # handle Unicode objects.
        writer.writerow([unicode(s).encode(encoding) for s in row])

    response = Response()
    response.body = stream.getvalue()
    stream.close()

    # Set Response headers.
    response.charset = encoding
    response.content_type = 'text/csv'
    response.content_disposition = 'attachment; filename={0}'.format(filename)

    return response


def xls_response(rows, encoding='ISO-8859-1', filename='response.xls', num_formats=(), col_widths=()):
    """Creates a downloadable webob.Response instance with the given row data
    as an XLS (Excel) file.

    :param rows: The spreadsheet data as rows.
    :type rows: iterable

    :param encoding: The character encoding for the data in the spreadsheet.
    :type encoding: str

    :param filename: The suggested filename for the user to save the file as.
    :type filename: str

    :param num_formats: A tuple of Excel format strings for numeric values.
    :type num_formats: tuple
    """
    stream = StringIO()

    wb = xlwt.Workbook(encoding=encoding)
    ws = wb.add_sheet('Nuorisovaalit 2011')

    # Compile the style objects
    styles = []
    for fmt in num_formats:
        if fmt is None:
            styles.append(None)
        else:
            styles.append(xlwt.easyxf(num_format_str=fmt))

    # Set column formatting styles.
    if len(rows) > 0 and len(num_formats) == len(rows[0]):
        for c in xrange(len(rows[0])):
            if styles[c] is not None:
                ws.col(c).set_style(styles[c])

    # Set column widths.
    if len(rows) > 0 and len(col_widths) == len(rows[0]):
        for c in xrange(len(rows[0])):
            if col_widths[c] is not None:
                ws.col(c).width = col_widths[c]

    for r, row in enumerate(rows):
        for c, item in enumerate(row):
            if len(styles) == len(row) and styles[c] is not None:
                ws.write(r, c, item, styles[c])
            else:
                ws.write(r, c, item)

    wb.save(stream)

    response = Response()
    response.body = stream.getvalue()
    stream.close()

    # Set Response headers.
    response.content_type = 'application/vnd.ms-excel'
    response.content_type_params = {}
    response.content_disposition = 'attachment; filename={0}'.format(filename)

    return response


def xls_response_multiple(sheets, encoding='ISO-8859-1', filename='response.xls', num_formats=(), col_widths=()):
    """Creates a downloadable webob.Response instance with the given row data
    as an XLS (Excel) file.

    :param sheets: The spreadsheet data as an iterable of (sheet_name,
                   rows) tuples with the row data of a sheet in the
                   rows iterable.
    :type sheets: iterable

    :param encoding: The character encoding for the data in the spreadsheet.
    :type encoding: str

    :param filename: The suggested filename for the user to save the file as.
    :type filename: str

    :param num_formats: A tuple of Excel format strings for numeric values.
    :type num_formats: tuple
    """
    wb = xlwt.Workbook(encoding=encoding)

    # Compile the style objects
    styles = []
    for fmt in num_formats:
        if fmt is None:
            styles.append(None)
        else:
            styles.append(xlwt.easyxf(num_format_str=fmt))

    if not sheets:
        wb.add_sheet(u'Tyhjä taulukko')

    for sheet_name, rows in sheets:

        # Worksheet name must not be longer than 31 characters.
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]

        # Create a new sheet and write to it.
        ws = wb.add_sheet(sheet_name)

        # Set column widths.
        if len(rows) > 0 and len(col_widths) == len(rows[0]):
            for c in xrange(len(rows[0])):
                if col_widths[c] is not None:
                    ws.col(c).width = col_widths[c]

        for r, row in enumerate(rows):
            for c, item in enumerate(row):
                if len(styles) == len(row) and styles[c] is not None:
                    ws.write(r, c, item, styles[c])
                else:
                    ws.write(r, c, item)

    stream = StringIO()
    wb.save(stream)

    response = Response()
    response.body = stream.getvalue()
    stream.close()

    # Set Response headers.
    response.content_type = 'application/vnd.ms-excel'
    response.content_type_params = {}
    response.content_disposition = 'attachment; filename={0}'.format(filename)

    return response


class SubmitVoters(object):
    """Receives a submission for the voter information corresponding to the
    current user's school.

    The latest submission (if available) is always rendered and the user may
    resubmit the information as many times as required. Only the latest
    submission is persisted overwriting any previous submission.

    The submission is validated during form processing.

    The submission functionality is available only during a configured time
    period and will be disabled otherwise.
    """
    #: Names of the fields in parsed CSV document.
    CSV_FIELDS = ('firstname', 'lastname', 'dob', 'gsm', 'email', 'address')
    #: Human readable names in the CSV document source.
    CSV_HEADER = (u'Etunimet (pakollinen)', u'Sukunimi (pakollinen)', u'Syntymäaika (pakollinen)', u'GSM-numero', u'Sähköpostiosoite', u'Kotiosoite')
    #: Excel field formatters. "@" means a text field.
    CSV_FORMATS = ('@', '@', 'DD.MM.YYYY', '@', '@', '@')

    def __init__(self, request):
        self.request = request
        assert len(SubmitVoters.CSV_FIELDS) == len(SubmitVoters.CSV_HEADER)
        assert len(SubmitVoters.CSV_FIELDS) == len(SubmitVoters.CSV_FORMATS)

    def submit_voters(self):
        """Handles the form submission and renders the currently persisted
        information.
        """
        user = authenticated_user(self.request)
        log = logging.getLogger('nuorisovaalitadmin')
        if user is None:
            log.warn('Unauthenticated attempt to submit voters.')
            raise Forbidden

        session = DBSession()
        options = {
            'errors': None,
            'errors_left': 0,
            'file_type_error': None,
        }

        if 'csvfile' in self.request.POST and hasattr(self.request.POST['csvfile'], 'file'):
            if self.request.POST.get('csrf_token') != self.request.session.get_csrf_token():
                log.warn('CSRF attempt at: {0}.'.format(self.request.url))
                raise Forbidden

            csv = self.request.POST['csvfile']
            errors = []

            try:
                try:
                    # Attempt to parse it as an Excel file first.
                    submission = self.parse_xls(csv.file)
                    log.info('Received Excel submission from "{0}".'.format(user.username))
                except xlrd.XLRDError:
                    # Fall back to CSV format.
                    csv.file.seek(0)
                    submission = self.parse_csv(csv.file)
                    log.info('Received CSV submission from "{0}".'.format(user.username))

                errors = self.validate(submission)
            except:
                log.warn('User "{0}" attempted to submit a non compliant file.'.format(user.username))
                options['file_type_error'] = True

            if options['file_type_error'] is None and len(errors) == 0:
                # Clear submissions for the user's school.
                session.query(CSVSubmission).filter_by(kind=CSVSubmission.VOTER,
                                                       school_id=user.school.id).delete()
                # Add the latest submission.
                session.add(CSVSubmission(
                    self.normalize(submission),
                    user.school,
                    CSVSubmission.VOTER))
                self.request.session.flash(u'Tiedosto tallennettu onnistuneesti.')

                # Redirect to avoid form resubmission on page reload.
                return HTTPFound(location=route_url('submit_voters', self.request))
            else:
                log.warn('Validation failed for voter submission made by "{0}"'.format(user.username))
                self.request.session.flash(u'Lähetetyssä tiedostossa oli virheitä. '
                                           u'Korjaa virheet ja lähetä uudelleen.')
                errors_to_show = 15
                options['errors'] = errors[:errors_to_show]
                if len(errors) > errors_to_show:
                    options['errors_left'] = len(errors) - errors_to_show

                # Print out the validation errors for easier failure tracking.
                for error_msg in options['errors']:
                    log.warn(u'Validation error (L{lineno}): {msg}'.format(**error_msg))

        options.update({
            'csrf_token': self.request.session.get_csrf_token(),
            'submission': None,
            'template_url': route_url('voters_template_xls', self.request),
            'action_url': route_url('submit_voters', self.request),
            'data_length_text': u'0 henkilöä',
            })

        submission = session.query(CSVSubmission)\
                        .filter_by(kind=CSVSubmission.VOTER, school_id=user.school.id)\
                        .first()
        if submission is not None:
            data_length = len(submission.csv)
            options['data_length_text'] = u'1 henkilö' if data_length == 1 else u'{0} henkilöä'.format(data_length)
            options['submission'] = {
                'timestamp': submission.timestamp.strftime('%d.%m.%Y %H:%M'),
                'data': submission.csv,
            }

        self.request.add_response_callback(disable_caching)

        return options

    def normalize(self, submission):
        """Normalizes the already validated submission data by removing
        empty rows and unifying the date representation.
        """

        def g():
            for row in submission:
                if not row['dob'].strip():
                    # Empty row, skip it.
                    continue
                for rexp in RE_DOB_SHORT:
                    dob = rexp.match(row['dob'])
                    if dob is not None:
                        # Convert a short date to the full form.
                        year, mon, day = int(dob.group(3)), int(dob.group(2)), int(dob.group(1))
                        row['dob'] = date(year + 1900, mon, day).strftime('%d.%m.%Y')
                        break

                yield row

        return tuple(g())

    def parse_xls(self, xlsfile, encoding='latin-1'):
        """Parses the given XLS file into a tuple of dictionaries.

        The dictionary keys are given in :py:attr:`SubmitVoters.CSV_FIELDS`.

        :param xlsfile: A file-like object that contains the Excel data.
        :type xlsfile: file

        :param encoding: The XLS data encoding. Defaults to "latin-1".
        :type encoding: str

        :rtype: tuple
        """
        wb = xlrd.open_workbook(file_contents=xlsfile.read())
        ws = wb.sheet_by_index(0)
        log = logging.getLogger('nuorisovaalitadmin')
        submission = []

        # Iterate over the rows, skipping the first row which contains the
        # header.
        for row in xrange(1, ws.nrows):
            # Create a record which contains all the required fields by default.
            entry = dict((k, u'') for k in SubmitVoters.CSV_FIELDS)

            for col, (title, fmt) in enumerate(zip(SubmitVoters.CSV_FIELDS, SubmitVoters.CSV_FORMATS)):
                try:
                    value = ws.cell_value(row, col)
                except IndexError:
                    log.warn('The submitted file is missing a column: (R{0}, C{1}).'.format(row + 1, col + 1))
                    raise

                if fmt == '@':
                    # This should be a text type field.
                    if not isinstance(value, basestring):
                        # The value was incorrectly parsed as something other
                        # than text, most likely as a (float) numeric value.
                        # Because we were expecting text and our schema does
                        # not contain floats the best we can do is convert the
                        # number first to int and then to string.
                        log.warn('Expected string value, got "{0}".'.format(value))
                        value = unicode(int(value))

                    entry[title] = unicode(value).strip()
                else:
                    # This should be a date field. It will throw an exception
                    # if the value is incorrect. We convert the date to string
                    # form now and let the validator re-parse it.
                    if isinstance(value, float):
                        entry[title] = unicode(date(*xlrd.xldate_as_tuple(value, wb.datemode)[:3]).strftime('%d.%m.%Y'))
                    else:
                        # We don't have a float value for a date field. We just use
                        # the value as is and let validation take care
                        # of the rest.
                        log.warn('Expected date (float) value, got "{0}".'.format(value.strip()))
                        entry[title] = unicode(value).strip()

            submission.append(entry)

        return tuple(submission)

    def parse_csv(self, csvfile, encoding='latin-1'):
        """Parses the given CSV file into a tuple of dictionaries.

        The dictionary keys are given in :py:attr:`SubmitVoters.CSV_FIELDS`.

        :param csvfile: A file-like object that contains the CSV data.
        :type csvfile: file

        :param encoding: The CSV file encoding. Defaults to "latin-1".
        :type encoding: str

        :rtype: tuple
        """

        def sniff_delimiter(csvfile, delimiters=(',', ';')):
            """Attempts to sniff the delimiter used in the CSV file."""
            try:
                sample = csvfile.read(2048).splitlines()[1]
                num_fields = len(SubmitVoters.CSV_FIELDS)
                for delimiter in delimiters:
                    if len(sample.split(delimiter, num_fields - 1)) >= num_fields:
                        log = logging.getLogger('nuorisovaalitadmin')
                        log.info('Using CSV delimiter "{0}".'.format(delimiter))
                        return delimiter
            finally:
                # Make sure we always rewind the file.
                csvfile.seek(0)

            raise ValueError('Failed to sniff CSV delimiter')

        reader = csv.DictReader(csvfile, SubmitVoters.CSV_FIELDS, restkey=None, restval='', delimiter=sniff_delimiter(csvfile))
        submission = []

        # Consume the header row.
        reader.next()

        for row in reader:
            # Get rid of any extra fields under restkey.
            row.pop(None, None)
            submission.append(
                dict((k, v.decode(encoding).strip()) for k, v in row.iteritems()))

        return tuple(submission)

    def validate(self, submission):
        """Validates the submitted data and returns a list of error messages
        corresponding to validation failures.

        Each error message is a dictionary containing the line number where
        the error occurred and a message describing the error, e.g.::

            { 'lineno': 13, 'msg': u'Etunimi puuttuu.' }

        :param submission: Iterable of dictionary containing the submission
            data.
        :type submission: iterable

        :rtype: list
        """
        errors = []

        def missing(row, field):
            return row.get(field) is None or len(row[field].strip()) == 0

        def error(lineno, msg):
            return dict(lineno=lineno, msg=msg)

        def normalize_gsm(gsm):
            """Remove whitespace and hyphens from the GSM number."""
            return ''.join(gsm.strip().replace('-', '').split())

        # Start line numbering from two since the header row was
        # consumed in parse_csv.
        for lineno, row in enumerate(submission, start=2):
            if not any(v.strip() for v in row.values()):
                # Skip empty rows.
                continue

            if missing(row, 'firstname'):
                errors.append(error(lineno, u'Etunimi puuttuu.'))
            if missing(row, 'lastname'):
                errors.append(error(lineno, u'Sukunimi puuttuu.'))

            if missing(row, 'dob'):
                errors.append(error(lineno, u'Syntymäaika puuttuu.'))
            else:
                dob = RE_DOB.match(row['dob'])
                if dob is None:
                    for rexp in RE_DOB_SHORT:
                        dob = rexp.match(row['dob'])
                        if dob is not None:
                            break

                if dob is None:
                    errors.append(error(lineno, u'Syntymäaika "{dob}" on väärän muotoinen.'.format(**row)))
                else:
                    try:
                        year, mon, day = int(dob.group(3)), int(dob.group(2)), int(dob.group(1))
                        if year < 100:
                            # The year was specified in two digit form. For valid
                            # submission we only accept people born before 2000 so
                            # we can safely ignore them now and add 1900 to the year.
                            year += 1900
                        d = date(year, mon, day)
                        age = date.today().year - d.year
                        if  age < 12 or age > 25:
                            errors.append(error(lineno, u'Äänestäjän ikä pitää olla välillä 12-25.'))
                    except ValueError:
                        errors.append(error(lineno, u'Syntymäaika "{dob}" on väärän muotoinen.'.format(**row)))

            # At least one contact info field must be found.
            if missing(row, 'gsm') and missing(row, 'email') and missing(row, 'address'):
                errors.append(error(lineno, u'Ainakin yksi yhteystieto (GSM-numero, email, osoite) on pakollinen.'))

            # GSM-number is optional.
            if not missing(row, 'gsm') and RE_GSM.match(normalize_gsm(row['gsm'])) is None:
                errors.append(error(lineno, u'GSM-numero "{gsm}" on väärän muotoinen.'.format(**row)))

            # Email address is optional.
            if not missing(row, 'email') and RE_EMAIL.match(row['email']) is None:
                errors.append(error(lineno, u'Sähköpostiosoite "{email}" on väärän muotoinen.'.format(**row)))

        return errors


def voters_template_csv(request):
    """Downloads the CSV template file for voter information submission.

    The CSV template is static and same for all school representatives.
    """
    request.add_response_callback(disable_caching_explorer)
    return csv_response([SubmitVoters.CSV_HEADER], filename='nuorisovaalit2011-oppilastiedot.csv')


def voters_template_xls(request):
    """Downloads the Excel template file for voter information submission.

    The Excel template is static and same for all school representatives.
    """
    request.add_response_callback(disable_caching_explorer)
    return xls_response(
        [SubmitVoters.CSV_HEADER],
        filename='nuorisovaalit2011-oppilastiedot.xls',
        num_formats=SubmitVoters.CSV_FORMATS)


def voter_list(request):
    """Renders a page for downloading the school specific list of voters.
    """
    user = authenticated_user(request)
    if user is None:
        raise Forbidden

    options = {
        'title': u'Lista äänestäneistä',
        'school': user.school.name,
        'district': user.school.district.name,
        'voter_list_xls': route_url('voter_list_xls', request),
    }

    request.add_response_callback(disable_caching)
    return options


def voter_list_xls(request):
    """Downloads the school specific list of voters as an Excel sheet.

    The list contains all voters registered for the school and voters who have
    already cast their vote will have a timestamp noting the time the vote
    was cast. The timestamp (or lack thereof) can be used to either grant or
    deny access to the paper ballot.

    The Excel sheet contains the following information about each voter:

        * Last name
        * First name
        * Date of birth (dd.mm.yyyy)
        * Voting timestamp
    """
    user = authenticated_user(request)
    if user is None:
        raise Forbidden

    rows = [
        (u'Sukunimi', u'Etunimi', u'Syntymäaika', u'Äänestänyt'),
    ]
    session = DBSession()

    # Query all the voters and the voting timestamp if they have voted.
    voters = session.query(Voter.lastname, Voter.firstname, Voter.dob, VotingLog.timestamp.label('timestamp'))\
             .outerjoin(VotingLog)\
             .filter(Voter.school == user.school)\
             .order_by(Voter.lastname, Voter.firstname)

    for voter in voters:
        timestamp = ''
        if voter.timestamp is not None:
            timestamp = datetime.fromtimestamp(voter.timestamp).strftime('%d.%m.%Y %H:%M:%S')
        rows.append((voter.lastname,
                     voter.firstname,
                     unicode(voter.dob.strftime('%d.%m.%Y')),
                     unicode(timestamp)))

    request.add_response_callback(disable_caching_explorer)
    return xls_response(rows, filename='nuorisovaalit2011-aanestajalista.xls',
                        num_formats=('@', '@', '@', '@'),
                        col_widths=(7000, 7000, 3500, 7000))


class SubmitResults(object):
    """Receives a submission of the paper ballot results for the current
    user's school.

    The latest submission (if available) is always rendered and the user may
    resubmit the results as many times as required. Only the latest submission
    is persisted.

    The submission is validated against the following criteria:

        * Correct candidate numbers
        * Valid vote count

    The submission functionality is available only during a configured time
    period and will be disabled otherwise.
    """
    #: Names of the fields in the parsed data dictionary.
    XLS_FIELDS = ('number', 'name', 'votes')
    #: Human readable column names in the Excel sheet.
    XLS_HEADER = (u'Numero', u'Nimi', u'Äänimäärä')
    #: Excel field formatters. "@" means a text field, None the default numeric field.
    XLS_FORMATS = (None, '@', None)

    def __init__(self, request):
        self.request = request
        self.user = None

    def submit_results(self):
        """Handles the form submission and renders the currently persisted
        information.
        """
        user = authenticated_user(self.request)
        log = logging.getLogger('nuorisovaalitadmin')

        if user is None:
            raise Forbidden
        self.user = user

        session = DBSession()
        options = {
            'title': u'Tulostietojen lähetys',
            'errors': None,
            'errors_left': 0,
            'file_type_error': None,
        }

        if 'xlsfile' in self.request.POST and hasattr(self.request.POST['xlsfile'], 'file'):
            if self.request.POST['csrf_token'] != self.request.session.get_csrf_token():
                log.warn('CSRF attempt at: {0}.'.format(self.request.url))
                raise Forbidden

            xls = self.request.POST['xlsfile']
            errors = []

            try:
                try:
                    # Attempt to parse it as an Excel 97/XP/2000 or older first.
                    submission = self.parse_xls(xls.file)
                    log.info('Validating Excel submission from "{0}".'.format(user.username))
                except xlrd.XLRDError:
                    # Attempt to parse it as an Excel 2010 file.
                    xls.file.seek(0)
                    submission = self.parse_xlsx(xls.file)
                    log.info('Validating Excel 2010 (xlsx) submission from "{0}".'.format(user.username))

                errors = self.validate(submission)
            except:
                log.warn('User "{0}" attempted to submit a non compliant file.'.format(user.username))
                options['file_type_error'] = True

            if options['file_type_error'] is None and len(errors) == 0:
                # Clear submissions for the user's school.
                session.query(CSVSubmission)\
                    .filter_by(kind=CSVSubmission.RESULT, school_id=user.school.id)\
                    .delete()
                # Add the latest submission.
                session.add(CSVSubmission(self.normalize(submission), user.school, CSVSubmission.RESULT))
                self.request.session.flash(u'Tiedosto tallennettu onnistuneesti.')

                # Redirect to avoid form resubmission on page reload.
                return HTTPFound(location=route_url('submit_results', self.request))
            else:
                log.warn('Validation failed for results submission made by "{0}"'.format(user.username))
                self.request.session.flash(u'Lähetetyssä tiedostossa oli virheitä. '
                                           u'Korjaa virheet ja lähetä uudelleen.')
                errors_to_show = 15
                options['errors'] = errors[:errors_to_show]
                if len(errors) > errors_to_show:
                    options['errors_left'] = len(errors) - errors_to_show

        options.update({
            'csrf_token': self.request.session.get_csrf_token(),
            'submission': None,
            'template_url': route_url('results_template_xls', self.request),
            'action_url': route_url('submit_results', self.request),
        })

        submission = session.query(CSVSubmission)\
                     .filter_by(kind=CSVSubmission.RESULT, school_id=user.school.id)\
                     .first()
        if submission is not None:
            options['submission'] = {
                'timestamp': submission.timestamp.strftime('%d.%m.%Y %H:%M'),
                'data': submission.csv,
            }

        self.request.add_response_callback(disable_caching)
        return options

    def normalize(self, submission):
        """Normalizes the already validated submission data by removing
        empty rows and parsing integers.
        """

        def g():
            for row in submission:
                if not row['number'].strip():
                    # Empty row, skip it.
                    continue
                row['number'] = int(row['number'].strip())
                row['votes'] = int(row['votes'].strip())

                yield row

        return tuple(g())

    def parse_xlsx(self, xlsfile):
        """Parses the given .xlsx file into a tuple of dictionaries.

        The dictionary keys are given in :py:attr:`SubmitResults.XLS_FIELDS`.

        :param xlsfile: A file-like object that contains the Excel data.
        :type xlsfile: file

        :rtype: tuple
        """
        tempxls = tempfile.NamedTemporaryFile(delete=True)
        log = logging.getLogger('webidentity')
        submission = []

        with tempxls.file as f:
            # openpyxl requires a path to read the .xlsx file. We use a named
            # tempory file which gets deleted when the file is closed.
            f.write(xlsfile.read())
            f.flush()
            os.fsync(f.fileno())

            wb = load_workbook(filename=tempxls.name)
            ws = wb.get_active_sheet()
            for i, row in enumerate(ws.rows[1:], start=1):
                entry = dict((k, u'') for k in SubmitResults.XLS_FIELDS)
                if len(row) != len(SubmitResults.XLS_FIELDS):
                    log.warn('The submitted file is missing a column on row {0}.'.format(i))
                    raise IndexError

                for cell, title, fmt in zip(row, SubmitResults.XLS_FIELDS, SubmitResults.XLS_FORMATS):
                    value = cell.value

                    # Simply convert all values to unicode and let validation
                    # worry about them.
                    if isinstance(value, basestring):
                        value = cell.value.strip()
                    else:
                        # We only expect integer values so we convert all numeric
                        # value unconditionally.
                        value = unicode(int(cell.value)).strip()

                    entry[title] = value

                submission.append(entry)

        return tuple(submission)

    def parse_xls(self, xlsfile):
        """Parses the given XLS file into a tuple of dictionaries.

        The dictionary keys are given in :py:attr:`SubmitResults.XLS_FIELDS`.

        :param xlsfile: A file-like object that contains the Excel data.
        :type xlsfile: file

        :rtype: tuple
        """
        wb = xlrd.open_workbook(file_contents=xlsfile.read())
        ws = wb.sheet_by_index(0)
        log = logging.getLogger('nuorisovaalitadmin')
        submission = []

        # Iterate over the rows, skipping the first row which contains the
        # header.
        for row in xrange(1, ws.nrows):
            # Create a record which contains all the required fields by default.
            entry = dict((k, u'') for k in SubmitResults.XLS_FIELDS)

            for col, (title, fmt) in enumerate(zip(SubmitResults.XLS_FIELDS, SubmitResults.XLS_FORMATS)):
                try:
                    value = ws.cell_value(row, col)
                except IndexError:
                    log.warn('The submitted file is missing a column: (R{0}, C{1}).'.format(row + 1, col + 1))
                    raise

                # Simply convert all values to unicode and let validation
                # worry about them.
                if isinstance(value, basestring):
                    value = value.strip()
                else:
                    # We only expect integer values so we convert all numeric
                    # value unconditionally.
                    value = unicode(int(value)).strip()

                entry[title] = unicode(value).strip()
            submission.append(entry)

        return tuple(submission)

    def validate(self, submission):
        """Validates the submitted data and returns a list of error messages
        corresponding to validation failures.

        Each error message is a dictionary containing the line number where
        the error occurred and a message describing the error, e.g.::

            { 'lineno': 13, 'msg': u'Etunimi puuttuu.' }

        :param submission: Iterable of dictionary containing the submission
            data.
        :type submission: iterable

        :rtype: list
        """
        errors = []

        def missing(row, field):
            return row.get(field) is None or len(row[field].strip()) == 0

        def error(lineno, msg):
            return dict(lineno=lineno, msg=msg)

        def is_num(n):
            return n is not None and n.strip().isdigit()

        # Start line numbering from two since the header row was
        # consumed while parsing.
        for lineno, row in enumerate(submission, start=2):
            if not any(v.strip() for v in row.values()):
                # Skip empty rows.
                continue

            if missing(row, 'name'):
                errors.append(error(lineno, u'Ehdokkaan nimi puuttuu.'))

            if missing(row, 'votes'):
                errors.append(error(lineno, u'Ehdokkaan äänimäärä puuttuu.'))
            elif not is_num(row.get('votes')):
                errors.append(error(lineno, u'Äänimäärä on väärää muotoa. Tarkista, että äänimäärä on positiivinen kokonaisluku.'))

            if missing(row, 'number'):
                errors.append(error(lineno, u'Ehdokasnumero puuttuu.'))
            elif not is_num(row.get('number')):
                errors.append(error(lineno, u'Ehdokasnumero on väärää muotoa. Tarkista, että ehdokasnumero on positiivinen kokonaisluku.'))
            else:
                # Check that there exists exactly one candidate with
                # the given number who is in the same district as the
                # user.
                num = int(row.get('number').strip())
                session = DBSession()
                num_candidates = session.query(Candidate)\
                                 .filter(Candidate.number == num)\
                                 .filter(Candidate.district == self.user.school.district)\
                                 .count()
                if num_candidates != 1:
                    errors.append(error(lineno, u'Ehdokasnumero {0} ei vastaa yhtään vaalipiirin ehdokasta.'.format(num)))

        return errors


def results_template_xls(request):
    """Downloads an Excel sheet template for paper ballot results submission.

    The Excel template contains the names and numbers of the candidates
    associated with the school's voting district.
    """
    rows = [
        (u'Numero', u'Nimi', u'Äänimäärä'),
    ]

    user = authenticated_user(request)
    if user is None:
        raise Forbidden

    session = DBSession()

    # Add all the candidates and their numbers into the CSV template.
    candidates = session.query(Candidate)\
                 .filter(Candidate.district == user.school.district)\
                 .order_by(Candidate.number)
    rows.extend((c.number, c.fullname(), u'') for c in candidates)

    request.add_response_callback(disable_caching_explorer)
    return xls_response(rows,
                        filename='nuorisovaalit2011-uurnatulokset.xls',
                        num_formats=('@', '@', '@'),
                        col_widths=(None, 8000, 3000))


def results(request):
    """Renders a page for downloading the school specific combined results of
    the election (electronic and paper ballot).
    """
    user = authenticated_user(request)
    if user is None:
        raise Forbidden
    options = {
        'title': u'Tulokset',
        'district': user.school.district.name,
        'results_template_xls': route_url('results_template_xls', request),
        'results_xls': route_url('results_xls', request),
    }

    request.add_response_callback(disable_caching)
    return options


def results_xls(request):
    """Downloads the district specific combined results as an Excel sheet.
    """
    user = authenticated_user(request)
    if user is None:
        raise Forbidden

    district = user.school.district

    title = u'{0} ({1} kansanedustajapaikka'.format(district.name, district.quota)
    title += ')' if district.quota == 1 else 'a)'
    rows = [
        (title, u'', u'', u'', u''),
        (u'', u'', u'', u'', u''),
        (u'Valitut ehdokkaat', u'', u'', u'', u''),
        (u'Numero', u'Nimi', u'Puolue', u'Äänimäärä', u'Vertailuluku'),
    ]

    for i, record in enumerate(dhont_selection(user.school.district)):

        # Add a separator between the selected and other candidates.
        if i == district.quota:
            rows.append((u'', u'', u'', u'', u''))
            rows.append((u'Valitsematta jääneet ehdokkaat', u'', u'', u'', u''))
            rows.append((u'Numero', u'Nimi', u'Puolue', u'Äänimäärä', u'Vertailuluku'))

        candidate = record['candidate']
        rows.append((candidate.number,
                     candidate.fullname(),
                     candidate.party.name,
                     record['absolute_votes'],
                     record['proportional_votes']))

    request.add_response_callback(disable_caching_explorer)
    return xls_response(rows,
                        filename='nuorisovaalit2011-tulokset.xls',
                        num_formats=(None, '@', '@', None, None),
                        col_widths=(None, 8000, 14000, 3000, 3000))
